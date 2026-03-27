#!/usr/bin/env python3
"""
Generate xlsx files from transcript data, one per final stage.

Can be used standalone with batch JSON files or imported as a library.

Standalone usage:
    python3 gen_xlsx.py /tmp/batch_0.json /tmp/batch_200.json ...
"""
import json
import os
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

S3_BUCKET = "reevo-prod-ng-meeting-transcripts-bucket"
S3_PREFIX = f"s3://{S3_BUCKET}/"


def _slugify(name):
    """Convert stage name to filename-safe slug: 'Closed Won' -> 'closed_won'."""
    return re.sub(r"[^a-z0-9]+", "_", name.strip().lower()).strip("_")


def write_xlsx(rows, path, sheet_title):
    """Write rows to an xlsx file. Returns row count."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [
        "Opportunity Name",
        "Stage Transition Date",
        "Meeting Date",
        "Transcript S3 URL",
        "Real Transcript S3 URL",
    ]
    hfont = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    dfont = Font(name="Arial", size=10)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row["opp"]).font = dfont
        ws.cell(row=i, column=2, value=row["neg_date"]).font = dfont
        ws.cell(row=i, column=3, value=row["mtg_date"]).font = dfont
        ws.cell(row=i, column=4, value=row["s3_url"]).font = dfont
        ws.cell(row=i, column=5, value=row["real_url"]).font = dfont

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 100
    ws.column_dimensions["E"].width = 100
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"

    wb.save(path)
    return len(rows)


def generate_xlsx_files(rows, output_dir):
    """Group rows by stage and write one xlsx per stage. Returns {stage: (path, count)}."""
    os.makedirs(output_dir, exist_ok=True)

    # Add S3 URL fields to each row
    enriched = []
    for r in rows:
        enriched.append({
            **r,
            "s3_url": S3_PREFIX + r["s3_key"],
            "real_url": S3_PREFIX + r["s3_key"],
        })

    # Group by stage
    stages = {}
    for r in enriched:
        stages.setdefault(r["stage"], []).append(r)

    results = {}
    for stage_name, stage_rows in sorted(stages.items()):
        slug = _slugify(stage_name)
        filename = f"{slug}_transcripts.xlsx"
        path = os.path.join(output_dir, filename)
        count = write_xlsx(stage_rows, path, stage_name)
        results[stage_name] = (path, count)
        print(f"  Wrote {count} rows to {path}")

    return results


# --- Standalone batch-JSON mode (backwards compatible) ---

def load_batches(filepaths):
    all_rows = []
    for f in filepaths:
        with open(f) as fh:
            data = json.load(fh)
            for row in data:
                all_rows.append({
                    "stage": row[0],
                    "opp": row[1],
                    "neg_date": row[2],
                    "mtg_date": row[3],
                    "s3_url": S3_PREFIX + row[4],
                    "real_url": S3_PREFIX + row[4],
                })
    return all_rows


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 gen_xlsx.py batch1.json batch2.json ...")
        sys.exit(1)

    output_dir = os.environ.get("OUTPUT_DIR", "/Users/akash/Desktop/transcripts")
    os.makedirs(output_dir, exist_ok=True)

    all_rows = load_batches(sys.argv[1:])

    # Group by stage
    stages = {}
    for r in all_rows:
        stages.setdefault(r["stage"], []).append(r)

    print(f"Total rows: {len(all_rows)}")
    for stage_name, stage_rows in sorted(stages.items()):
        slug = _slugify(stage_name)
        filename = f"{slug}_transcripts.xlsx"
        path = os.path.join(output_dir, filename)
        count = write_xlsx(stage_rows, path, stage_name)
        print(f"  {stage_name}: {count} rows -> {path}")


if __name__ == "__main__":
    main()
