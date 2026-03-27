#!/usr/bin/env python3
"""
Download transcript JSON files from S3 using xlsx files as input.

Can be used standalone or imported as a library.

Standalone usage:
    python3 download_transcripts.py --output-dir ./output --stages "Closed Won" "Closed Lost"
"""
import os
import re
import sys
import argparse

import boto3
from openpyxl import load_workbook

UNSAFE_CHARS = re.compile(r'[/\\:*?"<>|]')


def sanitize(name):
    return UNSAFE_CHARS.sub("_", name).strip()


def _slugify(name):
    """Convert stage name to filename-safe slug: 'Closed Won' -> 'closed_won'."""
    return re.sub(r"[^a-z0-9]+", "_", name.strip().lower()).strip("_")


def parse_s3_url(url):
    url = url.strip()
    if not url.startswith("s3://"):
        raise ValueError(f"Not an S3 URL: {url}")
    without_prefix = url[5:]
    bucket, _, key = without_prefix.partition("/")
    return bucket, key


def read_xlsx(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        opp_name = row[0]
        mtg_date = row[2]
        s3_url = row[4]
        if not s3_url:
            continue
        rows.append({
            "opp": str(opp_name),
            "date": str(mtg_date),
            "s3_url": str(s3_url),
        })
    wb.close()
    return rows


def build_filename(opp, date, seen):
    base = f"{sanitize(opp)}_{date}"
    key = base.lower()
    count = seen.get(key, 0)
    seen[key] = count + 1
    if count == 0:
        return f"{base}.json"
    return f"{base}_{count}.json"


def download_stage(rows, folder, label, s3):
    """Download all transcripts for a single stage. Returns (downloaded, skipped, failed)."""
    os.makedirs(folder, exist_ok=True)
    seen = {}
    downloaded = 0
    skipped = 0
    failed = []
    total = len(rows)

    for i, r in enumerate(rows, 1):
        fname = build_filename(r["opp"], r["date"], seen)
        dest = os.path.join(folder, fname)

        if os.path.exists(dest):
            skipped += 1
            continue

        try:
            bucket, key = parse_s3_url(r["s3_url"])
            print(f"  [{i}/{total}] {fname}")
            s3.download_file(bucket, key, dest)
            downloaded += 1
        except Exception as e:
            failed.append((fname, str(e)))
            print(f"  [{i}/{total}] FAILED {fname}: {e}")

    print(f"\n{label}: {downloaded} downloaded, {skipped} skipped (exist), {len(failed)} failed")
    return downloaded, skipped, failed


def download_all_transcripts(output_dir, final_stages):
    """Iterate stages, find xlsx files, create folders, download transcripts."""
    s3 = boto3.client("s3")
    summary = {}

    for stage_name in final_stages:
        slug = _slugify(stage_name)
        xlsx_path = os.path.join(output_dir, f"{slug}_transcripts.xlsx")
        folder = os.path.join(output_dir, f"{stage_name} Transcripts")

        if not os.path.exists(xlsx_path):
            print(f"WARNING: {xlsx_path} not found, skipping {stage_name}")
            continue

        print(f"\n=== {stage_name} Transcripts ===")
        rows = read_xlsx(xlsx_path)
        dl, sk, fail = download_stage(rows, folder, stage_name, s3)
        summary[stage_name] = {"downloaded": dl, "skipped": sk, "failed": fail, "total": len(rows)}

    return summary


# --- Standalone mode ---

def main():
    parser = argparse.ArgumentParser(description="Download transcript JSON files from S3")
    parser.add_argument("--output-dir", default="/Users/akash/Desktop/transcripts", help="Directory containing xlsx files")
    parser.add_argument("--stages", nargs="+", default=["Closed Won", "Closed Lost"], help="Stage names to download")
    args = parser.parse_args()

    summary = download_all_transcripts(args.output_dir, args.stages)

    print("\n=== Summary ===")
    all_failed = []
    for stage_name, s in summary.items():
        print(f"{stage_name}: {s['downloaded']} downloaded, {s['skipped']} skipped, {len(s['failed'])} failed (of {s['total']} total)")
        all_failed.extend(s["failed"])

    if all_failed:
        print("\nFailed downloads:")
        for fname, err in all_failed:
            print(f"  {fname}: {err}")
        sys.exit(1)


if __name__ == "__main__":
    main()
